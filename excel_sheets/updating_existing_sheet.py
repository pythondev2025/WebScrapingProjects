import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font


# loading workbook
wb = openpyxl.load_workbook("census_result.xlsx")
sheet = wb.active

# making dict of updated values of tracts
TRACTS_UPDATES = {
    "Bibb": 15,
    "Butler": 20,
    "Douglas": 100,
    "Weston": 5
}

# changing the values
for rownum in range(2, sheet.max_row + 1):
    county_name = sheet.cell(rownum, 2).value
    if county_name in TRACTS_UPDATES:
        sheet.cell(rownum, 4).value = TRACTS_UPDATES[county_name]

# applying the font
max_row = sheet.max_row
max_col = sheet.max_column
max_col = get_column_letter(max_col)
# sheet["A1":max_col + str(max_row)]
sheet_font = Font(name= "Calibri", size=10, bold=True, italic= True, color="0000FF")
for row in sheet:
    for cell in row:
        cell.font = sheet_font

# using the formula to calculate the total
sheet.cell(row=sheet.max_row + 1, column=3).value = "=SUM(C2:C" + str(sheet.max_row) + ")"
sheet.cell(row=sheet.max_row, column=4).value = "=SUM(D2:D3144)"
sheet["A" + str(sheet.max_row)] = "Total"

# freezing the panes
sheet.sheet_view.topLeftCell = "A1"
sheet.freeze_panes = "B2"

# saving updated copy of workbook
wb.save("updated_census_results.xlsx")
