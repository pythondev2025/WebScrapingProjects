import openpyxl

# creating wb
wb = openpyxl.Workbook()
sheet = wb.active
print(wb.sheetnames)
print(sheet.title)

# changing sheet title
sheet.title = "new_excel_sheet"
print(sheet.title)

# creating new sheet
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet()

# creating sheet with kwargs
wb.create_sheet("2nd_sheet", 1)
del wb["Sheet"]
print(wb.sheetnames)

# writing in sheet
sheet["a2"] = "Assalam o Alaikum"

# changing the dimensions of the cells
sheet.row_dimensions[3].height = 50
sheet.column_dimensions["B"].width = 30

# merging cells
sheet.merge_cells("A3:D6")
sheet["a3"] = "merged cells"
sheet.merge_cells("F6:H8")

# unmerge cells
sheet.unmerge_cells("F6:H8")
sheet["F6"] = "unmerged cells"

# freezing panes
sheet.freeze_panes = "B1"

wb.save("sample_sheet.xlsx")
