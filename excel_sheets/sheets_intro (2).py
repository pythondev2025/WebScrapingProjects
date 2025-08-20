import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook("example.xlsx")
sheet = wb["Sheet1"]


def intro(sheet):
    print(wb.sheetnames)
    name = sheet.title
    print(name)
    sheet = wb.active
    print(sheet)
    print(sheet.title)


# cell = sheet["B2"]
# print(cell.value)
# print(cell.coordinate)
# print(cell.row)
# print(cell.column)
# cell2 = sheet["A1"]
# print(cell2.value)

def easy_way():
    rows = sheet.max_row
    cols = sheet.max_column
    table = ''

    for i in range(1, rows+1):
        table += f"{i} "
        for j in range(1, cols+1):
            if j == cols:
                table += str(sheet.cell(i, j).value) + "\n"
            else:
                table += str(sheet.cell(i, j).value) + " "
    # print(table)
    with open("table.txt", "w") as file:
        file.write(table)
    x = column_index_from_string("zzz")
    print(x)
    y = get_column_letter(313)
    print(y)
    # when we iterate the sheet over the for loop its eah row would be iterated and when again iterate
    # each row over the for loop each cell class will be iterated
def slicing():
    for rows in sheet["a2":"c5"]:
        for cell in rows:
            print(cell.value)
    for col in sheet.columns:
        for cell in col:
            print(cell.value)
    for cell in list(sheet.rows)[0]:
        print(cell.value)


